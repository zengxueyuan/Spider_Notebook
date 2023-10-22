# file = open("./tmp/data.csv", "w")
# file.write("哪吒,姜子牙,花木兰")
# file.close()


# import csv
#
# with open(r"./tmp/book.csv", "w", encoding="utf-8", newline="") as file:
#     writer = csv.writer(file)
#     writer.writerow(["斗破苍穹", "天蚕土豆"])


# import csv
#
# with open(r"./tmp/book.csv", "r", encoding="utf-8", newline="") as file:
#     reader = csv.reader(file)
#     for row in reader:
#         print(row)


# import requests
# import bs4
# import csv
#
# res = requests.get("https://www.suning.com")
#
# with open("./tmp/suning.html", "wb") as file:
#     file.write(res.content)
#
# html_file = open("./tmp/suning.html", encoding="utf-8")
# soup = bs4.BeautifulSoup(html_file, "lxml")
#
# with open(r"./tmp/商品分类.csv", "w", encoding="utf-8", newline="") as file:
#     writer = csv.writer(file)
#     goods_list = []
#     tag_list = soup.select(".index-list a")
#     for tag in tag_list:
#         goods_list.append(tag.getText())
#     writer.writerow(goods_list)


# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = "new_title"
#
# wb.save(r"./tmp/first.xlsx")


# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = "happy"
#
# wb.save(r"./tmp/my_excel.xlsx")


# import openpyxl
#
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = "new_title"
#
# sheet['A1'].value = "酷我音乐"
# sheet['B2'].value = "汪苏泷"
#
# wb.save(r"./tmp/first.xlsx")


# import openpyxl
# wb = openpyxl.load_workbook("./tmp/first.xlsx")
# sheet = wb['new_title']
# print(sheet['A1'].value)


# import openpyxl
#
# wb = openpyxl.load_workbook("./tmp/my_excel.xlsx")
#
# sheet_list = wb.sheetnames
#
# for sheet_name in sheet_list:
#     sheet = wb[sheet_name]
#
#     for i in range(1, sheet.max_row + 1):
#         print(sheet['A' + str(i)].value)


import requests
import time
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
row_num = 1

url = "https://comment.kuwo.cn/com.s"

# 定义请求参数
my_params = {
    "type": "get_comment",
    "f": "web",
    "rows": "20",
    "digest": "2",
    "sid": "93",
    "uid": "0",
    "prod": "newWeb",
    "httpsStatus": 1,
    "reqId": "de3bd9d0-5da7-11ee-be91-09b4fcdf5e69",
    "plat": "web_www",
}

my_headers = {
    "Referer": "https://www.kuwo.cn/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36 Edg/117.0.2045.43"
}

for page in range(3):
    print("===爬取第{}页===".format(page + 1))
    my_params['page'] = page + 1
    res = requests.get(url, params=my_params, headers=my_headers)

    res_dict = res.json()
    comments = res_dict["rows"]

    for comment in comments:
        sheet['A' + str(row_num)].value = comment['msg']
        row_num += 1

    time.sleep(2)

wb.save("./tmp/酷我音乐评论.xlsx")
print("保存成功")
